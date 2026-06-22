"""
sync_core.py
------------
Pure data logic for syncing the MASTER sheet, kept separate from the
Streamlit UI so it can be unit-tested on its own.

The rule (from the spec):
  * SOURCE  = the Google-Sheet MASTER  -> supplies every OPERATIONAL column
  * TARGET  = a local Excel workbook   -> supplies the 7 FINANCIAL columns
  * OUTPUT  = a merged MASTER, matched on (DATE, STATION):
              operational values come from SOURCE,
              the 7 financial values are preserved from TARGET.
"""

from __future__ import annotations

import io
import re
from copy import copy

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- The 7 columns the source must NOT touch (kept/owned by the target) ----
FINANCIAL_COLS = [
    "BANKED",
    "Banking Date",
    "Bank",
    "Amount Deposited",
    "Balance Left",
    "PMS Cost",
    "AGO Cost",
]

# Columns used to line up a source row with a target row.
KEY_COLS = ["DATE", "STATION"]

# In this workbook the real column names live on row 2 (1-indexed),
# and row 1 holds the merged "group" headers (PRICE, TRUCKS, ...).
HEADER_ROW = 2          # 1-indexed row that holds the column names
GROUP_HEADER_ROW = 1    # 1-indexed row that holds the grouped banner headers
DEFAULT_SHEET = "MASTER"


# --------------------------------------------------------------------------- #
# Reading
# --------------------------------------------------------------------------- #
def extract_sheet_id(url: str) -> str:
    """Pull the spreadsheet id out of any Google-Sheets URL (or a bare id)."""
    url = (url or "").strip()
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    if re.fullmatch(r"[a-zA-Z0-9_-]{20,}", url):
        return url
    raise ValueError("Could not find a Google Sheet ID in that URL.")


def fetch_google_xlsx_bytes(url: str, creds_file: str | None = None) -> bytes:
    """
    Download a Google-hosted spreadsheet as raw .xlsx bytes.

    * No creds_file  -> uses the public /export?format=xlsx endpoint. The sheet
                        must be shared as 'Anyone with the link -> Viewer'.
    * creds_file set -> uses an authenticated Drive download (alt=media), which
                        works for *private* Drive-stored .xlsx files. Requires
                        `google-auth`; imported lazily so it's only needed when
                        you actually point at a service account.

    (Note: alt=media is for files uploaded to Drive as .xlsx — which is this
    workbook's case. A *native* Google Sheet would need ?export instead.)
    """
    sid = extract_sheet_id(url)

    if creds_file:
        from google.oauth2.service_account import Credentials
        import google.auth.transport.requests as gtr

        creds = Credentials.from_service_account_file(
            creds_file, scopes=["https://www.googleapis.com/auth/drive.readonly"]
        )
        creds.refresh(gtr.Request())
        api = f"https://www.googleapis.com/drive/v3/files/{sid}?alt=media"
        resp = requests.get(
            api, headers={"Authorization": f"Bearer {creds.token}"}, timeout=90
        )
        resp.raise_for_status()
        return resp.content

    export_url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"
    resp = requests.get(export_url, timeout=90)
    resp.raise_for_status()
    head = resp.content[:2000].lower()
    if b"<html" in head or b"<!doctype html" in head:
        raise RuntimeError(
            "Google returned a web page instead of a spreadsheet. The sheet is "
            "probably not shared publicly. Set link sharing to "
            "'Anyone with the link -> Viewer', or set GOOGLE_SERVICE_ACCOUNT_FILE "
            "in your .env to read it with a service account."
        )
    return resp.content


def read_master_from_bytes(
    data: bytes, sheet_name: str = DEFAULT_SHEET, header_row: int = HEADER_ROW
) -> pd.DataFrame:
    """Parse a MASTER tab out of in-memory .xlsx bytes."""
    return pd.read_excel(
        io.BytesIO(data), sheet_name=sheet_name, header=header_row - 1
    )


def read_google_master(
    url: str,
    sheet_name: str = DEFAULT_SHEET,
    header_row: int = HEADER_ROW,
    creds_file: str | None = None,
) -> pd.DataFrame:
    """Convenience: fetch a Google sheet and return the tab as a DataFrame."""
    return read_master_from_bytes(
        fetch_google_xlsx_bytes(url, creds_file), sheet_name, header_row
    )


def write_back_to_drive(url: str, data: bytes, creds_file: str) -> dict:
    """
    Replace the content of a Drive-stored .xlsx file *in place* with `data`.

    Keeps the same file id / share links — only the bytes change. Requires a
    service account that has been granted **Editor** access to the file, and the
    full drive scope. `google-auth` is imported lazily.

    Returns the Drive API response (id, name, modifiedTime, ...).
    """
    from google.oauth2.service_account import Credentials
    import google.auth.transport.requests as gtr

    sid = extract_sheet_id(url)
    creds = Credentials.from_service_account_file(
        creds_file, scopes=["https://www.googleapis.com/auth/drive"]
    )
    creds.refresh(gtr.Request())

    api = (
        f"https://www.googleapis.com/upload/drive/v3/files/{sid}"
        "?uploadType=media&fields=id,name,modifiedTime"
    )
    resp = requests.patch(
        api,
        headers={
            "Authorization": f"Bearer {creds.token}",
            "Content-Type": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        },
        data=data,
        timeout=180,
    )
    resp.raise_for_status()
    return resp.json()


def read_excel_master(
    file, sheet_name: str = DEFAULT_SHEET, header_row: int = HEADER_ROW
) -> pd.DataFrame:
    """Read a MASTER tab from an uploaded/loaded .xlsx file-like or path."""
    return pd.read_excel(file, sheet_name=sheet_name, header=header_row - 1)


# --------------------------------------------------------------------------- #
# Merging
# --------------------------------------------------------------------------- #
def _normalize_keys(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "DATE" in df.columns:
        df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce")
    if "STATION" in df.columns:
        df["STATION"] = df["STATION"].astype("string").str.strip()
    return df


def _valid_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop blank spacer rows: keep rows that have both a DATE and a STATION."""
    ok = df["DATE"].notna() & df["STATION"].notna() & (df["STATION"].str.len() > 0)
    return df[ok]


def build_synced_master(
    source_df: pd.DataFrame,
    target_df: pd.DataFrame,
    financial_cols: list[str] = FINANCIAL_COLS,
    key_cols: list[str] = KEY_COLS,
) -> tuple[pd.DataFrame, dict]:
    """
    Merge operational (source) + financial (target) into one MASTER frame.

    Returns (merged_df, report) where report has counts for the UI.
    Column order follows the SOURCE's original order.
    """
    final_order = list(source_df.columns)

    s = _valid_rows(_normalize_keys(source_df))
    t = _normalize_keys(target_df)
    t = t[t["DATE"].notna() & t["STATION"].notna()]

    # operational side = everything in the source except the financial columns
    op_cols = [c for c in s.columns if c not in financial_cols]
    s_op = s[op_cols].copy()

    # financial side = keys + whichever financial cols actually exist in target
    fin_present = [c for c in financial_cols if c in t.columns]
    t_fin = t[key_cols + fin_present].copy()
    # if a key appears twice in the target, keep the last non-empty entry
    t_fin = t_fin.drop_duplicates(subset=key_cols, keep="last")

    merged = s_op.merge(t_fin, on=key_cols, how="left", indicator=True)

    matched = int((merged["_merge"] == "both").sum())
    new_rows = int((merged["_merge"] == "left_only").sum())
    merged = merged.drop(columns="_merge")

    # make sure every original column exists, then restore original order
    for c in final_order:
        if c not in merged.columns:
            merged[c] = pd.NA
    merged = merged[final_order]

    # how many financial cells did we actually carry over?
    fin_filled = int(
        merged[fin_present].notna().sum().sum()
    ) if fin_present else 0

    report = {
        "source_rows": int(len(s)),
        "target_rows": int(len(t)),
        "output_rows": int(len(merged)),
        "matched_rows": matched,
        "new_rows_no_financial": new_rows,
        "financial_cols_found_in_target": fin_present,
        "financial_cells_preserved": fin_filled,
    }
    return merged, report


# --------------------------------------------------------------------------- #
# Writing back into a workbook (keeps every other tab intact)
# --------------------------------------------------------------------------- #
def write_synced_workbook(
    target_bytes: bytes,
    merged_df: pd.DataFrame,
    out_sheet_name: str = "MASTER_SYNCED",
    source_sheet_for_headers: str = DEFAULT_SHEET,
    group_header_row: int = GROUP_HEADER_ROW,
    overwrite_master: bool = False,
) -> bytes:
    """
    Take the original target workbook (as bytes), add/replace a sheet with the
    merged MASTER, and return the new workbook as bytes. All other tabs in the
    target workbook are left untouched.

    out_sheet_name : where to write the merged result.
    overwrite_master : if True, write into the existing MASTER sheet instead of
                       a new sheet (keeps the grouped banner + header rows).
    """
    wb = load_workbook(io.BytesIO(target_bytes))

    src_ws = wb[source_sheet_for_headers]
    # capture the grouped banner row (row 1) so the new sheet looks identical
    n_cols = len(merged_df.columns)
    banner = []
    for c in range(1, n_cols + 1):
        cell = src_ws.cell(row=group_header_row, column=c)
        banner.append((cell.value, copy(cell.font), copy(cell.fill),
                       copy(cell.alignment)))

    if overwrite_master:
        ws = wb[source_sheet_for_headers]
        # clear everything from the header row downward, keep rows 1..HEADER_ROW
        max_r = ws.max_row
        for r in range(HEADER_ROW + 1, max_r + 1):
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).value = None
        start_row = HEADER_ROW + 1
    else:
        if out_sheet_name in wb.sheetnames:
            del wb[out_sheet_name]
        ws = wb.create_sheet(out_sheet_name)
        # row 1: grouped banner
        for i, (val, font, fill, align) in enumerate(banner, start=1):
            cell = ws.cell(row=group_header_row, column=i, value=val)
            cell.font, cell.fill, cell.alignment = font, fill, align
        # row 2: column names
        for i, col in enumerate(merged_df.columns, start=1):
            ws.cell(row=HEADER_ROW, column=i, value=str(col))
        start_row = HEADER_ROW + 1

    # write the data
    date_cols = {i for i, c in enumerate(merged_df.columns, start=1)
                 if str(c).strip().upper() in ("DATE", "BANKING DATE")}
    for r_off, (_, row) in enumerate(merged_df.iterrows()):
        r = start_row + r_off
        for i, col in enumerate(merged_df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            cell = ws.cell(row=r, column=i, value=val)
            if i in date_cols and val is not None:
                cell.number_format = "yyyy-mm-dd"

    # widen the date column a touch
    for i in date_cols:
        ws.column_dimensions[get_column_letter(i)].width = 12

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
